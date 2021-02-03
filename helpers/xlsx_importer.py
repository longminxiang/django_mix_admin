import traceback
from openpyxl import load_workbook
from django.core.exceptions import ValidationError
from django.http import HttpResponseRedirect
from django.utils import timezone
from datetime import datetime


class HeaderIndex:

    def __init__(self, headers):
        super().__init__()
        for k, v in self.__class__.__dict__.items():
            if k.startswith('__'):
                continue
            idx = -1
            if isinstance(v, str):
                idx = headers.index(v) if v in headers else -1
            elif isinstance(v, (tuple, list)):
                for vv in v:
                    idx = headers.index(vv) if vv in headers else -1
                    if idx != -1:
                        break
            setattr(self, k, idx)


class XlsxFormatter:
    _pk = None
    _key_map = {}
    _key_index_map = None
    _did_get_headers = False
    _values = {}

    def __init__(self):
        super().__init__()
        for k, v in self.__class__.__dict__.items():
            if k.startswith('_') or not isinstance(v, (str, tuple, list)):
                continue
            self._key_map[k] = [v] if isinstance(v, str) else v

    def get_headers(self, vals):
        if self._did_get_headers:
            return True
        vals = [v.replace('\n', '') if isinstance(v, str) else v for v in vals]
        self._key_index_map = {}
        for k, indexs in self._key_map.items():
            idx = -1
            for index in indexs:
                idx = vals.index(index) if index in vals else -1
                if idx != -1:
                    break
            self._key_index_map[k] = idx
        if len([idx for idx in self._key_index_map.values() if idx != -1]) >= 3:
            self._did_get_headers = True
        return False

    def format(self, vals):
        hi = self.__class__()
        vals = [str(v).strip() if isinstance(v, str) else v for v in vals]
        self._values = {}
        for k, idx in self._key_index_map.items():
            val = vals[idx] if len(vals) > idx and idx >= 0 else None
            setattr(hi, k, val)
            hi._values[k] = val
        return hi if getattr(hi, self._pk, None) else None


class XlsxImporter:

    formatter_class = HeaderIndex
    error_messages = []
    message_prefix = None
    request = None

    def __init__(self, request):
        super().__init__()
        self.request = request
        self.error_messages = []

    def add_error(self, msg):
        self.error_messages.append((self.message_prefix or '') + msg)

    def _process(self, format_val):
        pass

    def _dispatched(self):
        pass

    def _pre_dispatch(self):
        pass

    def dispatch(self):
        f = self.request.FILES.get('file', None)
        if f is None:
            raise ValidationError('导入文件出错')
        if f.name.split('.')[-1] == 'xls':
            raise ValidationError('暂不支持xls格式，请另存为xlxs再导入')

        wb = load_workbook(f, data_only=True)
        ws = wb.worksheets[0]

        formatter = self.formatter_class()
        self._pre_dispatch()
        for vals in ws.iter_rows(values_only=True):
            vals = [v.strip() if isinstance(v, str) else v for v in vals]
            if not formatter.get_headers(vals):
                continue
            format_val = formatter.format(vals)
            if format_val is None:
                continue
            self.message_prefix = '【{}】：'.format(format_val.name)

            try:
                self._process(format_val)
            except Exception as e:
                print(e)
                print(traceback.format_exc(-3))
                self.add_error('存在错误')

            # 有错误及时抛出，12条一次
            if len(self.error_messages) >= 12:
                raise ValidationError(self.error_messages)

        # 在返回前再抛出错误
        if len(self.error_messages) > 0:
            raise ValidationError(self.error_messages)
        self._dispatched()

    def response(self):
        return HttpResponseRedirect(self.request.path)


def format_date(val):
    if isinstance(val, str):
        for format in ['%Y-%m-%d', '%Y-%m', '%Y年%m月', '%Y/%m/%d']:
            try:
                val = timezone.datetime.strptime(val, format)
                break
            except Exception:
                pass
    if isinstance(val, datetime):
        if timezone.is_naive(val):
            val = timezone.make_aware(val)
        return val.date()
